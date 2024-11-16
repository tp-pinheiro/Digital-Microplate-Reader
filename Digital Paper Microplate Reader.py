# -*- coding: utf-8 -*-
"""
Created on Wed Feb 28 11:41:32 2018

@author:Tomás Pinheiro - Tese Mestrado

Title - Funções para manipulação de imagem e extração de features
        Functions for image manipulation and feature extraction
"""

# IMPORTS
import numpy as np

#skimage imports
import skimage as ski
from skimage import io
from skimage import color
from skimage import filters
from skimage import measure
from skimage import exposure
from skimage import data
from skimage import morphology as mp
import matplotlib.pyplot as plt

#openpyxl imports
from openpyxl import *
import time


"""----------------------------------------------------------------------------------------"""

def get_labels(filename):
    """parameters: filename of plate picture"""
    """returns: labelled_image"""
    #open image from tif/jpeg/png file
    original_image= io.imread(filename)
    copy_image=original_image
    #gives cropping length according to image resolution
    resolution=copy_image.shape[0]
    print(resolution)
    crop_length = 25
    print(crop_length)
    print(int(resolution*0.57142857)-1000)
    print(int(resolution*0.285714)-500)
    # convert to greyscale
    grey = (color.rgb2grey(copy_image))

    #threshold with otsu method
    #note:in some cases, brightness manipulation might be needed
    thresh = filters.threshold_otsu(grey)
    binarized = ski.img_as_ubyte(grey > ((thresh)))

    plt.imshow(binarized);plt.show()
    #clean image through erosion
    #circle = mp.disk(15)
    #binarized = mp.erosion(binarized, selem=circle)
    #binarized = ski.img_as_ubyte(binarized < (thresh*255))

    #compute labelled image
    label_image, nb_labels = mp.label(binarized, return_num=True)
    
    return original_image, copy_image, label_image, crop_length, resolution

def get_centroids_coordenates_24(labelled_image, crop_length, resolution):
    """parameters: labelled image of 24 well microplate with ROIs"""
    """returns: coordenate vectors of centroid for ROIs """
    properties = measure.regionprops(labelled_image)
    centroid_list_hight = []
    centroid_list_width = []
    for n, values in enumerate(properties):
        if (properties[n].area < 50000 and properties[n].area > 1000):
            print(properties[n].area)
            if(properties[n].centroid[0] > crop_length and properties[n].centroid[1]>crop_length):
                centroid_list_hight += [properties[n].centroid[0]]
                centroid_list_width += [properties[n].centroid[1]]
    return centroid_list_width, centroid_list_hight

def get_centroids(labelled_image, crop_length, resolution, well_number):
    """parameters: labelled image of 24 well microplate with ROIs"""
    """returns: coordenate vectors of centroid for ROIs """
    properties = measure.regionprops(labelled_image)
    centroid_list_hight = []
    centroid_list_width = []
    for n, values in enumerate(properties):
        #if (properties[n].area < (int(resolution*0.57142857)-1000) and properties[n].area > (int(resolution*0.285714)-500)):
        if (properties[n].area < 3000 and properties[n].area > 2000):
             if (properties[n].centroid[0] > crop_length and properties[n].centroid[1] > crop_length):
                centroid_list_hight += [properties[n].centroid[0]]
                centroid_list_width += [properties[n].centroid[1]]
    minimum_hight = min(centroid_list_hight)
    maximum_hight = max(centroid_list_hight)
    minimum_width = min(centroid_list_width)
    maximum_width = max(centroid_list_width)
    centroid_hight=[]
    centroid_width=[]
    hight_difference = maximum_hight-minimum_hight
    if(well_number==96):
        hight_between_wells = int(hight_difference*0.1175)
        width_between_wells = int(hight_difference*0.1175)
        print(hight_between_wells)
        for i in range(1,9):
                set_width = minimum_width + int(0.158*hight_difference)
                set_hight = (minimum_hight + int(hight_difference*0.096))+((i-1)*hight_between_wells)
                for j in range(1,13):
                    centroid_width += [set_width]
                    centroid_hight += [set_hight]
                    set_width += width_between_wells
    if(well_number==24):
        hight_between_wells = int(hight_difference * 0.2396)
        width_between_wells = int(hight_difference * 0.2318)
        for i in range(1, 5):
            set_width = minimum_width + int(0.2357 * hight_difference)
            set_hight = (minimum_hight + int(hight_difference * 0.1589)) + ((i - 1) * hight_between_wells)
            for j in range(1, 7):
                centroid_width += [set_width]
                centroid_hight += [set_hight]
                set_width += width_between_wells

    return centroid_width, centroid_hight

def order_ROIs(centroid_list_width):
    ordered_width = []
    #for 24 well plates
    for i in range (1,5): 
        ordered_line=[]
        ordered_line = sorted(centroid_list_width[(6*i)-6:(6*i)])
        ordered_width += ordered_line      
    
    return ordered_width

def is_outlier(points, thresh):
    """
    Returns a boolean array with True if points are outliers and False 
    otherwise.

    Parameters:
    -----------
        points : An numobservations by numdimensions array of observations
        thresh : The modified z-score to use as a threshold. Observations with
            a modified z-score (based on the median absolute deviation) greater
            than this value will be classified as outliers.

    Returns:
    --------
        mask : A numobservations-length boolean array.

    References:
    ----------
        Boris Iglewicz and David Hoaglin (1993), "Volume 16: How to Detect and
        Handle Outliers", The ASQC Basic References in Quality Control:
        Statistical Techniques, Edward F. Mykytka, Ph.D., Editor. 
    """
    if len(points.shape) == 1:
        points = points[:,None]
    median = np.median(points, axis=0)
    diff = np.sum((points - median)**2, axis=-1)
    diff = np.sqrt(diff)
    med_abs_deviation = np.median(diff)

    modified_z_score = 0.6745 * diff / med_abs_deviation

    return modified_z_score > thresh

def equalization_image(image, hight, width, crop):
    crop = color.rgb2gray(image[int(hight[61])-crop:int(hight[61])+crop,int(width[61])-crop:int(width[61])+crop])
    #crop[:,:,1]=0
    #crop[:, :, 2] = 0
    img_eq = exposure.equalize_hist(crop)
    #img_norm = exposure.rescale_intensity(crop,  out_range=(0, 255))
    plt.imshow(img_eq, cmap=plt.cm.gray)
    plt.show()
    return 1

def get_RGB(original_image, centroid_list_hight, centroid_list_width, crop_length):
    """parameters: original image and lists of hight and width for ROIs"""
    """returns: mean R, G and B lists for each ROI  """
    #lists:
    mean_red = []
    mean_green = []
    mean_blue = []
    ms_red = []
    ms_blue = []
    frac = []
     
    for i, values in enumerate(centroid_list_hight):
        red_vector = []
        green_vector = []
        blue_vector = []
        fraction = []
        #crop original image:        
        cropped = original_image[int(centroid_list_hight[i]-crop_length):int(centroid_list_hight[i]+crop_length),int(centroid_list_width[i]-crop_length):int(centroid_list_width[i]+crop_length)]
        #retrieve features from pixels:
        for j in range(cropped.shape[0]):
            for m in range(cropped.shape[1]):
                red_vector += [cropped[j][m][0]]
                green_vector += [cropped[j][m][1]]
                blue_vector += [cropped[j][m][2]]
                #combination of features can be computed here
                #fraction += [(cropped[j][m][0])/(cropped[j][m][2])]
                #greycolour:
                #fraction += [int(np.mean(cropped[j][m]))]
        """
        #outlier_detection in pixels:
        thresh = 1
        red_outlier = is_outlier(np.asarray(red_vector), thresh)
        green_outlier = is_outlier(np.asarray(green_vector), thresh)
        blue_outlier = is_outlier(np.asarray(blue_vector), thresh)
        frac_outlier = is_outlier(np.asarray(fraction), thresh)
        final_red = []; final_green=[];final_blue =[]; #final_frac=[]
        for k in range(0, 1600):
            if(red_outlier[k]==False):
                final_red += [red_vector[k]]
            if(green_outlier[k]==False):
                final_green += [green_vector[k]]
            if(blue_outlier[k]==False):
                final_blue += [blue_vector[k]]
            #if(frac_outlier[k]==False):
                #final_frac += [fraction[k]]
        """

        #experiment------------------------------------------------------
        #histogram of pixel distribution       
        """
        if(i==61):
            plt.hist(red_vector, bins=range(0,255), color='red');plt.show()
            #print(np.mean(fraction));print(np.std(fraction))
            plt.hist(green_vector, bins=range(0,255), color='green');plt.show()
            #print(np.mean(fraction));print(np.std(fraction))
            plt.hist(blue_vector, bins=range(0,255), color='blue');plt.show()
           # print(np.mean(fraction));print(np.std(fraction))
        """

        #-----------------------------------------------------------------------  

        #mean values for colour features:
        mean_red += [np.mean(red_vector)] 
        mean_green += [np.mean(green_vector)]
        mean_blue += [np.mean(blue_vector)]
        frac += [np.mean(fraction)]

        """
        #most significant pixel-red
        a = np.histogram(red_vector, bins=range(min(red_vector), max(red_vector)))
        loc=np.where(a[0]==max(a[0]))
        ms_red += [a[1][int(np.mean(loc[0]))]]
        #most significant pixel_blue
        a = np.histogram(blue_vector, bins=range(min(blue_vector), max(blue_vector)))
        loc=np.where(a[0]==max(a[0]))
        ms_blue += [a[1][int(np.mean(loc[0]))]]
    """
    return mean_red, mean_green, mean_blue #, frac, ms_red, ms_blue

def get_HSV(original_image, centroid_list_hight, centroid_list_width, crop_length):
    """parameters: original image and lists of hight and width for ROIs"""
    """returns: mean H, S and V lists for each ROI  """
    h_vector = []
    s_vector = []
    v_vector = []
    for i, values in enumerate(centroid_list_hight):
        sum_h=0
        sum_s = 0
        sum_v = 0
        cropped = original_image[int(centroid_list_hight[i]-crop_length):int(centroid_list_hight[i]+crop_length),int(centroid_list_width[i]-crop_length):int(centroid_list_width[i]+crop_length)]
        #convertion to hsv colour space:
        hsv_crop = color.rgb2hsv(cropped)
        for j in range(hsv_crop.shape[0]):
            for m in range(hsv_crop.shape[1]):
                sum_h += hsv_crop[j][m][0]
                sum_s += hsv_crop[j][m][1]
                sum_v += hsv_crop[j][m][2]
            mean_h = sum_h/(cropped.shape[0]*cropped.shape[1])
            mean_s = sum_s/(cropped.shape[0]*cropped.shape[1])
            mean_v = sum_v/(cropped.shape[0]*cropped.shape[1])
        h_vector += [mean_h]
        s_vector += [mean_s]
        v_vector += [mean_v]
        
    return h_vector, s_vector, v_vector
    
def get_XYZ(original_image, centroid_list_hight, centroid_list_width, crop_length):
    """parameters: original image and lists of hight and width for ROIs"""
    """returns: mean X, Y and Z lists for each ROI  """
    x_vector = []
    y_vector = []
    z_vector = []
    for i, values in enumerate(centroid_list_hight):
        sum_x=0
        sum_y = 0
        sum_z = 0
        cropped = original_image[int(centroid_list_hight[i]-crop_length):int(centroid_list_hight[i]+crop_length),int(centroid_list_width[i]-crop_length):int(centroid_list_width[i]+crop_length)]
        xyz_crop = color.rgb2xyz(cropped) 
        for j in range(xyz_crop.shape[0]):
            for m in range(xyz_crop.shape[1]):
                sum_x += xyz_crop[j][m][0]
                sum_y += xyz_crop[j][m][1]
                sum_z += xyz_crop[j][m][2]
            mean_x = sum_x/(cropped.shape[0]*cropped.shape[1])
            mean_y = sum_y/(cropped.shape[0]*cropped.shape[1])
            mean_z = sum_z/(cropped.shape[0]*cropped.shape[1])
        x_vector += [mean_x]
        y_vector += [mean_y]
        z_vector += [mean_z]
        
    return x_vector, y_vector, z_vector
    
def plot_ROIs(original_image, hight_coordenates, width_coordenates):
    fig, ax = plt.subplots()
    ax.imshow(original_image)        
    #mark centroids from ROIs in the image
    for n, values in enumerate(hight_coordenates):
        ax.scatter(width_coordenates[n], hight_coordenates[n], marker = "$ {} $".format(n))
        
    plt.axis('off')
    #plt.xlim(0,image.shape[0])
    #plt.ylim((image.shape[1]/2),0)
    plt.show()

def write_excel_file(plate_type):
    #write excel file with resulting mean feature values for manipulation
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Date:'
    now = time.strftime("%x")
    ws['B1'] = now
    ws['D1'] = 'Plate type:'
    ws['E1'] = plate_type
    ws['B3']='R'
    ws['C3']='G'
    ws['D3']='B'
    ws['G3']='H'
    ws['H3']='S'
    ws['I3']='V'
    ws['L3']='X'
    ws['M3']='Y'
    ws['N3']='Z'
    if(plate_type==24):
        for i in range(0, len(R)):
            sheet_index = 'A'+str(i+5)
            if(i<6):
                text = 'A'+str(i+1)
            if(i<12 and i>=6):
                text = 'B'+str(i-5)
            if(i<18 and i>=12):
                text = 'C'+str(i-11)
            if(i>=18):
                text='D'+str(i-17)
            ws[sheet_index]=text
        for i in range(0,len(R)):
            sheet_index_R = 'B'+str(i+5)
            ws[sheet_index_R]=R[i]        
            sheet_index_G = 'C'+str(i+5)
            ws[sheet_index_G]=G[i]
            sheet_index_B = 'D'+str(i+5)
            ws[sheet_index_B]=B[i]
            sheet_index_H = 'G'+str(i+5)
            ws[sheet_index_H]=H[i]
            sheet_index_S = 'H'+str(i+5)
            ws[sheet_index_S]=S[i]
            sheet_index_V = 'I'+str(i+5)
            ws[sheet_index_V]=V[i]
            sheet_index_X = 'L'+str(i+5)
            ws[sheet_index_X]=X[i]
            sheet_index_Y = 'M'+str(i+5)
            ws[sheet_index_Y]=Y[i]
            sheet_index_Z = 'N'+str(i+5)
            ws[sheet_index_Z]=Z[i]
    if(plate_type == 96):
        for i in range(0, len(R)):
            sheet_index = 'A'+str(i+5)
            if(i<12):
                text = 'A'+str(i+1)
            if(i<24 and i>=12):
                text = 'B'+str(i-11)
            if(i<36 and i>=24):
                text = 'C'+str(i-23)
            if(i<48 and i>=36):
                text='D'+str(i-35)
            if(i<60 and i>=48):
                text='E'+str(i-47)
            if(i<72 and i>=60):
                text='F'+str(i-59)
            if(i<84 and i>=72):
                text='G'+str(i-71)
            if(i>=84):
                text='H'+str(i-83)
            ws[sheet_index]=text
        for i in range(0,len(R)):
            sheet_index_R = 'B'+str(i+5)
            ws[sheet_index_R]=R[i]        
            sheet_index_G = 'C'+str(i+5)
            ws[sheet_index_G]=G[i]
            sheet_index_B = 'D'+str(i+5)
            ws[sheet_index_B]=B[i]
            sheet_index_H = 'G'+str(i+5)
            ws[sheet_index_H]=H[i]
            sheet_index_S = 'H'+str(i+5)
            ws[sheet_index_S]=S[i]
            sheet_index_V = 'I'+str(i+5)
            ws[sheet_index_V]=V[i]
            sheet_index_X = 'L'+str(i+5)
            ws[sheet_index_X]=X[i]
            sheet_index_Y = 'M'+str(i+5)
            ws[sheet_index_Y]=Y[i]
            sheet_index_Z = 'N'+str(i+5)
            ws[sheet_index_Z]=Z[i]        
    if(plate_type == 384):#TERMINAR
        for i in range(0, len(R)):
            sheet_index = 'A'+str(i+5)
            if(i<12):
                text = 'A'+str(i+1)
            if(i<24 and i>=12):
                text = 'B'+str(i-11)
            if(i<36 and i>=24):
                text = 'C'+str(i-23)
            if(i<48 and i>=36):
                text='D'+str(i-35)
            if(i<60 and i>=48):
                text='E'+str(i-47)
            if(i<72 and i>=60):
                text='F'+str(i-59)
            if(i<84 and i>=72):
                text='G'+str(i-71)
            if(i>=84):
                text='H'+str(i-83)
            ws[sheet_index] = text
        for i in range(0,len(R)):
            sheet_index_R = 'B'+str(i+5)
            ws[sheet_index_R]=R[i]        
            sheet_index_G = 'C'+str(i+5)
            ws[sheet_index_G]=G[i]
            sheet_index_B = 'D'+str(i+5)
            ws[sheet_index_B]=B[i]
            sheet_index_H = 'G'+str(i+5)
            ws[sheet_index_H]=H[i]
            sheet_index_S = 'H'+str(i+5)
            ws[sheet_index_S]=S[i]
            sheet_index_V = 'I'+str(i+5)
            ws[sheet_index_V]=V[i]
            sheet_index_X = 'L'+str(i+5)
            ws[sheet_index_X]=X[i]
            sheet_index_Y = 'M'+str(i+5)
            ws[sheet_index_Y]=Y[i]
            sheet_index_Z = 'N'+str(i+5)
            ws[sheet_index_Z]=Z[i]        
        
    #insert directory for output excel file        
    wb.save("C:/Users/Tiago/Documents/18-19/CENIMAT/GlucoseSensorArticle/New_files/Lab/Interference/B.xlsx")

"""-------------------------------------------------------------------------------------------------------------------------------------------------"""    
#BEGINING OF USER CUSTOMIZATION - comment line when not needed

#insert image directory
image, copy, label_image, crop, resolution= get_labels("C:/Users/Tiago/Documents/18-19/CENIMAT/GlucoseSensorArticle/New_files/Lab/Interference/new-10.tif")
#width, hight = get_centroids_coordenates_24(label_image, crop, resolution)
width, hight = get_centroids(label_image, crop, resolution, 96)
#width = order_ROIs(width) #for 24 well plates

#96 well plate (uncomment in case 96 well plate is being used)
#width, hight = get_centroids_coordenates_96(label_image)

#384 well plate (uncomment in case 384 well plate is being used)
#width, hight = get_centroids_coordenates_384(label_image)

#plot image with marked ROIs
plot_ROIs(image, hight, width)

plt.imshow(image[int(hight[5])-crop:int(hight[5])+crop,int(width[5])-crop:int(width[5])+crop])
plt.show()
#equalization_image(image, hight, width, crop)
#get color space values
R, G, B = get_RGB(image, hight, width, crop)
H, S, V=get_HSV(image, hight, width, crop)
X, Y, Z = get_XYZ(image, hight, width, crop)

#write excel files with color space values (24,96,384)
write_excel_file(96)
print("DONE")
"""-------------------------------------------------------------------------------------------------------------------------------------------------"""